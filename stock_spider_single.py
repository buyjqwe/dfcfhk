import time
import random
import requests
import json
import openpyxl
from openpyxl import load_workbook
import sys
import os
import urllib3
import concurrent.futures
import datetime

# 尝试导入 setting
try:
    from setting import *
except ImportError:
    data = {"股票名称": "", "股票代码": "", "今开": "", "当前价": ""} 
    start_stock_code = 1
    end_stock_code = 100 
    read_last_code = 0
    file_path = f"StockData_{datetime.datetime.now().strftime('%Y-%m-%d')}.xlsx"

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
sys.path.append(os.path.dirname(__file__))

class stock_spider():
    def __init__(self):
        self.session = requests.Session()
        # 强制禁用代理
        self.session.trust_env = False 
        self.session.proxies = {'http': None, 'https': None}

        # 挂载适配器
        adapter = requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=20, max_retries=3)
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)

    def startExcel(self, formatted_data, file_path):
        directory = os.path.dirname(file_path)
        if directory and not os.path.exists(directory):
            os.makedirs(directory)

        try:
            self.wb = load_workbook(file_path)
            self.ws = self.wb.active
            print(f"成功加载 Excel: {file_path}")
        except FileNotFoundError:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Stock Data"
            print(f"创建新 Excel: {file_path}")
            
        if self.ws.max_row == 1:
            headers = list(formatted_data.keys())
            self.ws.append(headers)

    def insertExcel(self, formatted_data_list):
        for formatted_data in formatted_data_list:
            values = list(formatted_data.values())
            self.ws.append(values)

    def saveExcel(self, file_path):
        try:
            self.wb.save(file_path)
            print(f"--- 进度保存: 股票数据已更新到 {file_path} ---")
        except PermissionError:
            print(f"保存失败：请关闭 {file_path} 文件后再试！")

    def get_value(self, stock_data, key, factor=1, default="0", is_percentage=False, magnification=1):
        value = stock_data.get(key, "/")
        if value == "/" or value == '-' or value is None:
            return default
        if is_percentage:
            return f'{value / factor:.2f}%'
        if magnification != 1:
            value = value * magnification
        return str(value / factor if isinstance(value, (int, float)) else value)

    def formatId(self, id):
        return str(id).zfill(5)

    def convertToFloat(self, value):
        if value == "-":
            return 0
        else:
            try:
                value = f'{float(value):.3f}'
                return value
            except:
                return '0'

    def formatInfo(self, id, stock_data, financial_data):
        formatted_data = {
            "股票名称": stock_data["f58"],
            "股票代码": id,
            "今开": self.get_value(stock_data,"f46", 1000),
            "最高价": self.get_value(stock_data,"f44", 1000),
            "52周最高": self.get_value(stock_data,"f174", 1000),
            "成交量": self.get_value(stock_data,"f47", 10000, is_percentage=False),
            "外盘": self.get_value(stock_data,"f49", 10000),
            "昨收": self.get_value(stock_data,"f60", 1000),
            "最低价": self.get_value(stock_data,"f45", 1000),
            "52周最低": self.get_value(stock_data,"f175", 1000),
            "成交额": self.get_value(stock_data,"f48", 10000),
            "内盘": self.get_value(stock_data,"f161", 10000),
            "总股本": self.convertToFloat(self.get_value(stock_data,"f84", 100000000)),
            "港股本": self.convertToFloat(self.get_value(stock_data,"f85", 100000000)),
            "市净率": self.get_value(stock_data,"f167", 100),
            "每股收益": self.get_value(stock_data,"f108"),
            "股息率": self.get_value(stock_data,"f126", default="0", is_percentage=False),
            "总市值": self.convertToFloat(self.get_value(stock_data,"f116", 100000000)),
            "港市值": self.convertToFloat(self.get_value(stock_data,"f117", 100000000)),
            "市盈率": self.get_value(stock_data,"f164",magnification=0.01),
            "每股净资产": self.get_value(stock_data,"f92"),
            "换手率": self.get_value(stock_data,"f168", 100, is_percentage=True),
            "基本每股收益(元)": financial_data.get("BASIC_EPS"),
            "稀释每股收益(元)": financial_data.get("DILUTED_EPS"),
            "TTM每股收益(元)": financial_data.get("EPS_TTM"),
            "每股净资产(元)": financial_data.get("BPS"),
            "每股经营现金流(元)": financial_data.get("PER_NETCASH_OPERATE"),
            "每股营业收入(元)": financial_data.get("PER_OI"),
            "营业总收入(元)": self.convertToFloat(self.get_value(financial_data,"OPERATE_INCOME", 100000000)),
            "营业总收入同比增长(%)": financial_data.get("OPERATE_INCOME_YOY"),
            "营业总收入滚动环比增长(%)": financial_data.get("OPERATE_INCOME_QOQ"),
            "毛利润(元)": self.convertToFloat(self.get_value(financial_data,"GROSS_PROFIT", 100000000)),
            "毛利润同比增长(%)": financial_data.get("GROSS_PROFIT_YOY"),
            "毛利润滚动环比增长(%)": financial_data.get("GROSS_PROFIT_QOQ"),
            "归母净利润(元)": self.convertToFloat(self.get_value(financial_data,"HOLDER_PROFIT", 100000000)),
            "归母净利润同比增长(%)": financial_data.get("HOLDER_PROFIT_YOY"),
            "归母净利润滚动环比增长(%)": financial_data.get("HOLDER_PROFIT_QOQ"),
            "所得税/利润总额(%)": financial_data.get("TAX_EBT"),
            "经营现金流/营业收入(%)": financial_data.get("OCF_SALES"),
            "平均净资产收益率(%)": financial_data.get("ROE_AVG"),
            "年化净资产收益率(%)": financial_data.get("ROE_YEARLY"),
            "总资产净利率(%)": financial_data.get("ROA"),
            "毛利率(%)": financial_data.get("GROSS_PROFIT_RATIO"),
            "净利率(%)": financial_data.get("NET_PROFIT_RATIO"),
            "年化投资回报率(%)": financial_data.get("ROIC_YEARLY"),
            "应收账款周转率(次)": financial_data.get("ACCOUNTS_RECE_TDAYS"),
            "存货周转率(次)": financial_data.get("INVENTORY_TDAYS"),
            "流动资产周转率(次)": financial_data.get("CURRENT_ASSETS_TDAYS"),
            "总资产周转率(次)": financial_data.get("TOTAL_ASSETS_TDAYS"),
            "流动比率(倍)": financial_data.get("CURRENT_RATIO"),
            "流动负债/总负债(%)": financial_data.get("CURRENTDEBT_DEBT"),
            "资产负债率(%)": financial_data.get("DEBT_ASSET_RATIO"),
            "权益乘数": financial_data.get("EQUITY_MULTIPLIER"),
            "产权比率": financial_data.get("EQUITY_RATIO")
        }
        return formatted_data

    def get_financial_data(self, secucode, page_number=1, page_size=9):
        secucode += ".HK"
        url = (
            f"https://datacenter.eastmoney.com/securities/api/data/v1/get?"
            f"reportName=RPT_HKF10_FN_MAININDICATOR&columns=ALL&filter=(SECUCODE%3D%22{secucode}%22)&"
            f"pageNumber={page_number}&pageSize={page_size}&sortTypes=-1&sortColumns=STD_REPORT_DATE&"
            f"source=F10&client=PC&v=09194871783016205"
        )
        # 修改：移除Host，添加Referer，模拟真实访问来源
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            "Referer": "https://emweb.securities.eastmoney.com/"
        }
        
        for _ in range(3):
            try:
                # 修改：allow_redirects=False 禁止死循环，遇到302直接停止
                response = self.session.get(url, headers=headers, verify=False, timeout=10, allow_redirects=False)
                if response.status_code == 200:
                    json_result = response.json()
                    if json_result.get("result"):
                        return json_result["result"]["data"][0]
                    return {}
                # 如果是重定向，视为失败，不重试
            except Exception as e:
                pass 
            time.sleep(1)
        return {}

    def getStockInfo(self, id):
        url = "https://push2.eastmoney.com/api/qt/stock/get"
        params = {
            'invt': '2', 'fltt': '1', 'fields': 'f58,f107,f57,f43,f59,f169,f170,f152,f46,f60,f44,f45,f47,f48,f19,f532,f39,f161,f49,f171,f50,f86,f600,f601,f154,f84,f85,f168,f108,f116,f167,f164,f92,f71,f117,f177,f123,f124,f125,f174,f175,f126,f257,f256,f258,f251,f255,f252,f254,f253,f198,f292,f301,f752,f751',
            'secid': '116.' + str(id), 'ut': 'fa5fd1943c7b386f172d6893dbfba10b', 'dect': '1'
        }
        # 修改：移除Host，添加Referer
        headers = {
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
            "Referer": "https://quote.eastmoney.com/"
        }
        
        for _ in range(3):
            try:
                # 修改：allow_redirects=False
                response = self.session.get(url, headers=headers, params=params, verify=False, timeout=10, allow_redirects=False)
                if response.status_code == 200:
                    try:
                        return response.json().get("data")
                    except:
                        txt = response.text
                        if '(' in txt:
                            txt = txt[txt.index('(')+1 : txt.rindex(')')]
                            return json.loads(txt).get("data")
            except Exception:
                pass
            time.sleep(1)
        return None

    def read_last_stock_code(self, file_path):
        if not os.path.exists(file_path):
            return None
        try:
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            if ws.max_row > 1:
                return int(ws.cell(row=ws.max_row, column=2).value)
        except:
            return None
        return None

    def fetch_one_stock(self, id):
        time.sleep(random.uniform(0.5, 2)) 
        
        formatted_id = self.formatId(id)
        stock_data = self.getStockInfo(formatted_id)
        
        if stock_data is not None:
            financial_data = self.get_financial_data(formatted_id)
            formatted_data = self.formatInfo(formatted_id, stock_data, financial_data)
            return formatted_data
        return None

if __name__ == '__main__':
    start_time = time.time()
    s = stock_spider()
    
    print(f"Target Excel File: {file_path}")

    s.startExcel(data, file_path)
    
    start_code = start_stock_code
    
    if read_last_code == 1:
        last = s.read_last_stock_code(file_path)
        if last:
            start_code = last + 1
            print(f"Resuming from {start_code}")

    buffer_list = []
    save_interval = 50 
    max_workers = 8 

    print(f"Starting spider with {max_workers} workers...")
    
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        future_to_id = {executor.submit(s.fetch_one_stock, i): i for i in range(start_code, end_stock_code)}
        
        for future in concurrent.futures.as_completed(future_to_id):
            res = future.result()
            if res:
                buffer_list.append(res)
                print(f"Fetched: {res['股票代码']}")
            
            if len(buffer_list) >= save_interval:
                s.insertExcel(buffer_list)
                s.saveExcel(file_path)
                buffer_list = []

    if buffer_list:
        s.insertExcel(buffer_list)
        s.saveExcel(file_path)

    print(f"Done. Duration: {time.time() - start_time:.2f}s")
