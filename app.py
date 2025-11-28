import time
import random
import requests
import json
import openpyxl
from openpyxl import load_workbook
import sys
import os
import urllib3
import concurrent.futures # 引入线程池模块
# 导入 setting 配置 (确保你本地有 setting.py 文件)
from setting import *

# 禁用安全警告
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

sys.path.append(os.path.dirname(__file__))

class stock_spider():
    def __init__(self):
        # 初始化 Session，复用TCP连接，提高效率
        self.session = requests.Session()
        
        # === 关键修改：强制禁用代理 ===
        # 即使系统开了VPN，脚本也会直连，防止 ProxyError
        self.session.trust_env = False 
        self.session.proxies = {'http': None, 'https': None}

        # 挂载适配器，设置连接池大小
        adapter = requests.adapters.HTTPAdapter(pool_connections=20, pool_maxsize=20, max_retries=3)
        self.session.mount('http://', adapter)
        self.session.mount('https://', adapter)

    def startExcel(self, formatted_data, file_path):
        try:
            # 尝试加载已存在的 Excel 文件
            self.wb = load_workbook(file_path)
            self.ws = self.wb.active
            print(f"成功加载 Excel: {file_path}")
        except FileNotFoundError:
            # 如果文件不存在，则创建一个新的文件
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = "Stock Data"
            print(f"创建新 Excel: {file_path}")
            
        # 检查是否有表头（第1行），如果没有，则添加表头
        if self.ws.max_row == 1:
            headers = list(formatted_data.keys())
            self.ws.append(headers)  # 添加表头

    def insertExcel(self, formatted_data_list):
        # 批量添加数据
        for formatted_data in formatted_data_list:
            values = list(formatted_data.values())
            self.ws.append(values)

    def saveExcel(self, file_path):
        # 保存工作簿
        try:
            self.wb.save(file_path)
            print(f"--- 进度保存: 股票数据已更新到 {file_path} ---")
        except PermissionError:
            print(f"保存失败：请关闭 {file_path} 文件后再试！")

    # 定义一个函数，检查数据是否为空或无效
    def get_value(self, stock_data, key, factor=1, default="0", is_percentage=False, magnification=1):
        value = stock_data.get(key, "/")
        if value == "/" or value == '-' or value is None:
            return default
        # 如果是百分比格式，转换时添加百分号
        if is_percentage:
            return f'{value / factor:.2f}%'
        
        # 如果需要转换为单位
        if magnification != 1:
            value = value * magnification
        return str(value / factor if isinstance(value, (int, float)) else value)

    def formatId(self, id):
        return str(id).zfill(5)

    def convertToFloat(self, value):
        if value == "-":
            return 0
        else:
            try:
                value = f'{float(value):.3f}'
                return value
            except:
                return '0'

    def formatInfo(self, id, stock_data, financial_data):
        formatted_data = {
            "股票名称": stock_data["f58"],
            "股票代码": id,
            "今开": self.get_value(stock_data, "f46", 1000),  # f46 是今开，单位是分，转换为元
            "最高价": self.get_value(stock_data, "f44", 1000),  # f44 是最高价，单位是分
            "52周最高": self.get_value(stock_data, "f174", 1000),  # f174 是52周最高价
            "成交量": self.get_value(stock_data, "f47", 10000, is_percentage=False),  # f47 是成交量，单位是股，转换为万股
            "外盘": self.get_value(stock_data, "f49", 10000),  # f49 是外盘，单位是股，转换为万股
            "昨收": self.get_value(stock_data, "f60", 1000),  # f60 是昨收，单位是分
            "最低价": self.get_value(stock_data, "f45", 1000),  # f45 是最低价，单位是分
            "52周最低": self.get_value(stock_data, "f175", 1000),  # f175 是52周最低价
            "成交额": self.get_value(stock_data, "f48", 10000),  # f48 是成交额，单位是元，转换为万元
            "内盘": self.get_value(stock_data, "f161", 10000),  # f161 是内盘，单位是股，转换为万股
            "总股本": self.convertToFloat(self.get_value(stock_data, "f84", 100000000)),  # f84 是总股本，单位是股，转换为亿股
            "港股本": self.convertToFloat(self.get_value(stock_data, "f85", 100000000)),  # f85 是港股本，单位是股，转换为亿股
            "市净率": self.get_value(stock_data, "f167", 100),  # f167 是市净率
            "每股收益": self.get_value(stock_data, "f108"),  # f108 是每股收益
            "股息率": self.get_value(stock_data, "f126", default="0", is_percentage=False),  # f257 是股息率
            "总市值": self.convertToFloat(self.get_value(stock_data, "f116", 100000000)),  # f116 是总市值，单位是元，转换为亿元
            "港市值": self.convertToFloat(self.get_value(stock_data, "f117", 100000000)),  # f117 是港市值，单位是元，转换为亿元
            "市盈率": self.get_value(stock_data, "f164", magnification=0.01),  # f164 是市盈率
            "每股净资产": self.get_value(stock_data, "f92"),  # f92 是每股净资产
            "换手率": self.get_value(stock_data, "f168", 100, is_percentage=True),  # f168 是换手率
            "基本每股收益(元)": financial_data.get("BASIC_EPS"),
            "稀释每股收益(元)": financial_data.get("DILUTED_EPS"),
            "TTM每股收益(元)": financial_data.get("EPS_TTM"),
            "每股净资产(元)": financial_data.get("BPS"),
            "每股经营现金流(元)": financial_data.get("PER_NETCASH_OPERATE"),
            "每股营业收入(元)": financial_data.get("PER_OI"),
            "营业总收入(元)": self.convertToFloat(self.get_value(financial_data, "OPERATE_INCOME", 100000000)),  # 转换为亿
            "营业总收入同比增长(%)": financial_data.get("OPERATE_INCOME_YOY"),
            "营业总收入滚动环比增长(%)": financial_data.get("OPERATE_INCOME_QOQ"),
            "毛利润(元)": self.convertToFloat(self.get_value(financial_data, "GROSS_PROFIT", 100000000)),  # 转换为亿
            "毛利润同比增长(%)": financial_data.get("GROSS_PROFIT_YOY"),
            "毛利润滚动环比增长(%)": financial_data.get("GROSS_PROFIT_QOQ"),
            "归母净利润(元)": self.convertToFloat(self.get_value(financial_data, "HOLDER_PROFIT", 100000000)),  # 转换为亿
            "归母净利润同比增长(%)": financial_data.get("HOLDER_PROFIT_YOY"),
            "归母净利润滚动环比增长(%)": financial_data.get("HOLDER_PROFIT_QOQ"),
            "所得税/利润总额(%)": financial_data.get("TAX_EBT"),
            "经营现金流/营业收入(%)": financial_data.get("OCF_SALES"),
            "平均净资产收益率(%)": financial_data.get("ROE_AVG"),
            "年化净资产收益率(%)": financial_data.get("ROE_YEARLY"),
            "总资产净利率(%)": financial_data.get("ROA"),
            "毛利率(%)": financial_data.get("GROSS_PROFIT_RATIO"),
            "净利率(%)": financial_data.get("NET_PROFIT_RATIO"),
            "年化投资回报率(%)": financial_data.get("ROIC_YEARLY"),
            "应收账款周转率(次)": financial_data.get("ACCOUNTS_RECE_TDAYS"),
            "存货周转率(次)": financial_data.get("INVENTORY_TDAYS"),
            "流动资产周转率(次)": financial_data.get("CURRENT_ASSETS_TDAYS"),
            "总资产周转率(次)": financial_data.get("TOTAL_ASSETS_TDAYS"),
            "流动比率(倍)": financial_data.get("CURRENT_RATIO"),
            "流动负债/总负债(%)": financial_data.get("CURRENTDEBT_DEBT"),
            "资产负债率(%)": financial_data.get("DEBT_ASSET_RATIO"),
            "权益乘数": financial_data.get("EQUITY_MULTIPLIER"),
            "产权比率": financial_data.get("EQUITY_RATIO")
        }

        # print(f"get stockinfo {id}") # 减少打印，保持控制台整洁
        return formatted_data

    def get_financial_data(self, secucode, page_number=1, page_size=9):
        secucode += ".HK"
        url = (
            f"https://datacenter.eastmoney.com/securities/api/data/v1/get?"
            f"reportName=RPT_HKF10_FN_MAININDICATOR&"
            f"columns=ALL&"
            f"quoteColumns=&"
            f"filter=(SECUCODE%3D%22{secucode}%22)&"
            f"pageNumber={page_number}&"
            f"pageSize={page_size}&"
            f"sortTypes=-1&"
            f"sortColumns=STD_REPORT_DATE&"
            f"source=F10&"
            f"client=PC&"
            f"v=09194871783016205"
        )

        headers = {
            "Host": "datacenter.eastmoney.com",
            "Connection": "keep-alive", # 保持连接
            "sec-ch-ua-platform": "\"Windows\"",
            "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0",
            "sec-ch-ua": "\"Microsoft Edge\";v=\"129\", \"Not=A?Brand\";v=\"8\", \"Chromium\";v=\"129\"",
            "sec-ch-ua-mobile": "?0",
            "Accept": "*/*",
            "Origin": "https://emweb.securities.eastmoney.com",
            "Sec-Fetch-Site": "same-site",
            "Sec-Fetch-Mode": "cors",
            "Sec-Fetch-Dest": "empty",
            "Referer": "https://emweb.securities.eastmoney.com/",
            "Accept-Encoding": "gzip, deflate",
            "Accept-Language": "zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6"
        }

        response = None
        retry_count = 3  # 设置重试次数
        for _ in range(retry_count):
            try:
                # 使用 self.session 发送请求
                response = self.session.get(url, headers=headers, verify=False, timeout=10)
                if response.status_code == 200:
                    try:
                        json_result = response.json()
                        if json_result.get("result") is not None:
                            return json_result["result"]["data"][0]  # 返回JSON格式数据
                        else:
                            return {}  # 返回空字典
                    except json.JSONDecodeError:
                        print(f"JSON Decode Error for {secucode}")
                        return {}
                else:
                    print(f"Failed to fetch financial data {secucode}, status code: {response.status_code}, retrying...")
            except requests.exceptions.ConnectTimeout:
                print("Requests.exceptions.ConnectTimeout! retrying...")
            except requests.exceptions.RequestException as e:  # 捕获更广泛的 requests 异常
                print(f"Request Error: {e}, retrying...")
            
            time.sleep(1)  # 失败重试等待，减少到1秒
            response = None  # 重置

        return {"error": "Failed to fetch financial data after multiple retries."}

    def getStockInfo(self, id):
        # 定义请求URL和参数
        url = "https://push2.eastmoney.com/api/qt/stock/get"

        params = {
            'invt': '2',
            'fltt': '1',
            'cb': 'jQuery3510028574898925495473_1729038113447',
            'fields': 'f58,f107,f57,f43,f59,f169,f170,f152,f46,f60,f44,f45,f47,f48,f19,f532,f39,f161,f49,f171,f50,f86,f600,f601,f154,f84,f85,f168,f108,f116,f167,f164,f92,f71,f117,f177,f123,f124,f125,f174,f175,f126,f257,f256,f258,f251,f255,f252,f254,f253,f198,f292,f301,f752,f751',
            'secid': '116.' + str(id),
            'ut': 'fa5fd1943c7b386f172d6893dbfba10b',
            'wbp2u': '|0|0|0|web',
            'dect': '1',
            '_': '1729038113487'
        }

        # 定义请求头
        headers = {
            'Host': 'push2.eastmoney.com',
            'Connection': 'keep-alive', # 保持连接
            'Pragma': 'no-cache',
            'Cache-Control': 'no-cache',
            'sec-ch-ua-platform': '"Windows"',
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36 Edg/122.0.0.0',
            'sec-ch-ua': '"Microsoft Edge";v=\"129\", "Not=A?Brand";v=\"8\", "Chromium";v=\"129"',
            'sec-ch-ua-mobile': '?0',
            'Accept': '*/*',
            'Sec-Fetch-Site': 'same-site',
            'Sec-Fetch-Mode': 'no-cors',
            'Sec-Fetch-Dest': 'script',
            'Referer': 'https://quote.eastmoney.com/hk/03333.html',
            'Accept-Encoding': 'gzip, deflate',
            'Accept-Language': 'zh-CN,zh;q=0.9,en;q=0.8,en-GB;q=0.7,en-US;q=0.6'
        }

        # 发送请求
        response = None
        retry_count = 3  # 设置重试次数
        for _ in range(retry_count):
            try:
                # 使用 session 发送请求
                response = self.session.get(url, headers=headers, params=params, verify=False, timeout=10)
            except requests.exceptions.ConnectTimeout:
                print("Requests.exceptions.ConnectTimeout! retrying...")
            except requests.exceptions.RequestException as e:  # 捕获更广泛的 requests 异常
                print(f"Request Error: {e}, retrying...")
            
            if response and response.status_code == 200:
                jsonp_data = response.text
                try:
                    # 去除 JSONP 回调函数，提取有效的 JSON
                    if '(' in jsonp_data and ')' in jsonp_data:
                        financial_data = jsonp_data[jsonp_data.index('(') + 1: jsonp_data.rindex(')')]
                        data = json.loads(financial_data)
                        stock_data = data.get("data")
                        
                        if (stock_data is not None) and ('f58' in stock_data):
                            return stock_data
                    else:
                        print(f"Invalid JSONP format for {id}")
                except Exception as e:
                    print(f"Parsing Error for {id}: {e}")

            time.sleep(1)  # 失败重试等待
            response = None

        return None

    def read_last_stock_code(self, file_path):
        if not os.path.exists(file_path):
            print("文件不存在，将从头开始！")
            return None

        try:
            # 加载 Excel 文件
            wb = openpyxl.load_workbook(file_path)
            ws = wb.active
            # 检查第二列的表头（第1行，第2列），是否为"股票代码"
            if ws.cell(row=1, column=2).value != "股票代码":
                print("第二列的表头不是 '股票代码'！无法断点续传")
                return None
            
            # 读取最后一行第二列的数据（股票代码）
            if ws.max_row > 1:
                last_stock_code = ws.cell(row=ws.max_row, column=2).value
                return int(last_stock_code)
            else:
                return None
        except Exception as e:
            print(f"读取Excel失败: {e}")
            return None

    # 改名为 fetch_one_stock，并且去掉多线程逻辑，改为直接返回数据
    def fetch_one_stock(self, id):
        # 增加随机休眠时间，防止并发过快导致被限速
        time.sleep(random.uniform(2, 5)) 
        
        formatted_id = self.formatId(id)
        stock_data = self.getStockInfo(formatted_id)
        
        # 只有当获取到股票基础信息时才获取财务信息，节省请求
        if stock_data is not None:
            financial_data = self.get_financial_data(formatted_id)
            formatted_data = self.formatInfo(formatted_id, stock_data, financial_data)
            return formatted_data
        else:
            # print(f"Stock {id} not found or empty.")
            return None


if __name__ == '__main__':
    start_time = time.time()

    s = stock_spider()
    
    # 这里的 file_path 和 data 来自 import setting，请确保 setting.py 存在
    # 如果没有 setting.py，请手动在此处定义 file_path 和 data 字典
    
    s.startExcel(data, file_path)
    
    start_stock_code_effective = start_stock_code
    
    # 检查是否需要断点续传
    if read_last_code == 1:
        last_code = s.read_last_stock_code(file_path)
        if last_code:
            print(f"检测到上次抓取位置，从股票代码：{last_code + 1} 开始继续")
            start_stock_code_effective = last_code + 1
        else:
            print(f"未检测到有效记录，从配置的起始代码：{start_stock_code} 开始")
    
    buffer_list = [] # 缓存列表
    save_interval = 20 # 增加到每20条保存一次
    max_workers = 8 # 开启 8 个线程并发，既提速又比无限线程安全

    print(f"开始多线程抓取 (线程数: {max_workers})...")

    # 使用 ThreadPoolExecutor 管理并发
    with concurrent.futures.ThreadPoolExecutor(max_workers=max_workers) as executor:
        # 提交所有任务到线程池
        future_to_id = {executor.submit(s.fetch_one_stock, stock_id): stock_id for stock_id in range(start_stock_code_effective, end_stock_code)}
        
        # as_completed 会在某个任务完成时立刻让出控制权
        for future in concurrent.futures.as_completed(future_to_id):
            stock_id = future_to_id[future]
            try:
                result = future.result()
                if result:
                    buffer_list.append(result)
                    print(f"成功获取: {result['股票名称']} ({result['股票代码']})")
                else:
                    # print(f"跳过: {stock_id}")
                    pass
            except Exception as exc:
                print(f'Stock {stock_id} generated an exception: {exc}')

            # 批量写入和保存逻辑 (仍在主线程进行，是线程安全的)
            if len(buffer_list) >= save_interval:
                s.insertExcel(buffer_list)
                s.saveExcel(file_path)
                buffer_list = [] # 清空缓存

    # 循环结束后，保存剩余的数据
    if buffer_list:
        s.insertExcel(buffer_list)
        s.saveExcel(file_path)

    end_time = time.time()
    elapsed_time = end_time - start_time
    print(f"任务完成！总耗时: {elapsed_time:.2f} 秒")
